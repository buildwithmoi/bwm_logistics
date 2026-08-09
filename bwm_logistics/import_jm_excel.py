# Copyright (c) 2026, Build With Moi and contributors
# For license information, please see license.txt

"""The JM Containers working Excel → containers, trading shipments and
distribution entries.

Two ways in, both landing on the same records:

1. **Direct import** from a spreadsheet on this machine (needs openpyxl and the
   .xlsx to hand — client data is gitignored, so this is a local-only path):

       bench --site <site> execute bwm_logistics.import_jm_excel.run
       bench --site <site> execute bwm_logistics.import_jm_excel.run \\
           --kwargs "{'path': '/path/to.xlsx'}"

2. **Shipped opening balance.** `extract()` freezes a spreadsheet into
   `data/jm_containers_opening.json`, which is committed. The patch
   `bwm_logistics.patches.v1_0.import_client_opening_data` loads that file on
   `bench migrate`, so a fresh client site comes up with its real stock without
   anyone copying a spreadsheet onto the server:

       bench --site <site> execute bwm_logistics.import_jm_excel.extract \\
           --kwargs "{'path': '/path/to.xlsx'}"
       git add bwm_logistics/data/jm_containers_opening.json && git push

Both are idempotent: containers already imported (same container no + BL) are
skipped, as are identical distribution rows — safe to re-run after fixing data.
"""

import json
import os

import frappe
from frappe.utils import flt, getdate

from bwm_logistics.api.items import catalogue_blocker, ensure_item
from bwm_logistics.bwm_logistics.doctype.shipment.shipment import manifests_for

DEFAULT_FILENAME = "Stock and Distribution 1.xlsx"

# The frozen opening balance shipped in the repo (written by extract()).
OPENING_DATA = os.path.join("data", "jm_containers_opening.json")

# JM Containers operates a single branch.
BRANCH = "Accra"

# Their sheet statuses → our tracking milestones (which set shipment status).
STATUS_MILESTONE = {
	"Arrived": "Arrived at Port",
	"In Transit": "In Transit",
	"Pending": None,  # stays Open
	"Delayed": "Delayed",
	"Customs Clearance": "Customs Clearance",
	"Offloaded": "Offloaded",
}


def _norm(text) -> str:
	return (text or "").strip().lower()


def _default_path() -> str:
	# App repo root (…/apps/bwm_logistics), one level above the package dir.
	return os.path.abspath(os.path.join(frappe.get_app_path("bwm_logistics"), "..", DEFAULT_FILENAME))


def opening_data_path() -> str:
	return os.path.join(frappe.get_app_path("bwm_logistics"), OPENING_DATA)


def _cell(value):
	"""Spreadsheet cell → a clean string.

	Excel stores long numeric BLs as floats, so `2694851.0` has to come back as
	`"2694851"`, not `"2694851.0"`.
	"""
	if value is None:
		return None
	if isinstance(value, float) and value.is_integer():
		value = int(value)
	text = str(value).strip()
	return text or None


def _date(value):
	return str(getdate(value)) if value else None


# ── Parsing ─────────────────────────────────────────────────────────────────
def parse_workbook(path=None) -> dict:
	"""Read the workbook into the plain dict that both entry points consume.

	Doing the reading in one place means the committed JSON and a live .xlsx
	import can never drift apart.
	"""
	import openpyxl

	path = path or _default_path()
	if not os.path.exists(path):
		raise FileNotFoundError(path)

	wb = openpyxl.load_workbook(path, data_only=True)
	data = {"source": os.path.basename(path), "branch": BRANCH, "containers": [], "distributions": []}

	# ── Stock Tracker → one container + one trading shipment per row ─────────
	rows = list(wb["Stock Tracker"].iter_rows(values_only=True))
	header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Date Received")
	for r in rows[header_idx + 1 :]:
		(date_received, item, bl_no, container_no, eta, qty, unit, supplier, status, invoice_ref, comment) = (
			list(r) + [None] * 11
		)[:11]
		if not container_no or not item:
			continue

		# The Item column may hold several goods in one container ("A & B").
		# The sheet carries a single Qty per row, so it lands on the first line
		# and the rest start at 0, to be corrected on arrival.
		names = [p.strip() for p in str(item).replace("\n", " ").split("&") if p.strip()]
		packages = [
			{
				"description": name,
				"qty": int(flt(qty)) if (i == 0 and qty) else 0,
				"unit": (_cell(unit) or "PIECES").upper(),
			}
			for i, name in enumerate(names)
		]

		data["containers"].append(
			{
				"container_no": _cell(container_no),
				"bl_no": _cell(bl_no),
				"eta": _date(eta),
				"date_received": _date(date_received),
				"status": _cell(status),
				"supplier": _cell(supplier),
				"invoice_ref": _cell(invoice_ref),
				"comment": _cell(comment),
				"packages": packages,
				"qty_needs_split": len(names) > 1 and bool(qty),
			}
		)

	# ── Distribution → Distribution Entries ─────────────────────────────────
	rows = list(wb["Distribution"].iter_rows(values_only=True))
	header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Customer Name")
	for r in rows[header_idx + 1 :]:
		(recipient, product, qty, unit_price, _total, destination, delivery_date) = (list(r) + [None] * 7)[:7]
		if not recipient or not product or not flt(qty):
			continue
		data["distributions"].append(
			{
				"recipient": _cell(recipient),
				"product": _cell(product),
				"qty": flt(qty),
				"unit_price": flt(unit_price) or None,
				"destination": _cell(destination),
				"delivery_date": _date(delivery_date),
			}
		)

	return data


def extract(path=None, out=None):
	"""Freeze a spreadsheet into the committed opening-balance JSON."""
	data = parse_workbook(path)
	out = out or opening_data_path()
	os.makedirs(os.path.dirname(out), exist_ok=True)
	with open(out, "w") as f:
		json.dump(data, f, indent=2, ensure_ascii=False)
		f.write("\n")
	print(
		f"Wrote {out}: {len(data['containers'])} containers, "
		f"{len(data['distributions'])} distribution rows (source: {data['source']})"
	)
	return data


# ── Loading ─────────────────────────────────────────────────────────────────
def load(data: dict) -> dict:
	"""Create the records for a parsed workbook. Safe to re-run."""
	frappe.set_user("Administrator")
	frappe.flags.mute_emails = True

	# The manifest points at catalogue Items, which need ERPNext's Item Group
	# tree and UOM list — both built by the setup wizard, not by install. On a
	# server that has been provisioned but not set up, defer rather than fail:
	# normalise_opening_data_to_model_b retries on the next migrate.
	blocker = catalogue_blocker()
	if blocker:
		print(f"Opening import deferred — {blocker}. Re-run after setup:\n  bench --site <site> execute bwm_logistics.import_jm_excel.load_opening")
		return {"deferred": blocker}

	branch = data.get("branch") or BRANCH
	if not frappe.db.exists("Branch", branch):
		frappe.get_doc({"doctype": "Branch", "branch": branch}).insert(ignore_permissions=True)

	created = {"containers": 0, "shipments": 0, "events": 0, "distributions": 0}
	skipped = []
	shipments = []  # trading shipments in sheet order — distributions match against these

	for row in data["containers"]:
		container_no, bl_no = row["container_no"], row.get("bl_no")

		existing = frappe.db.get_value("Container", {"container_no": container_no, "bl_no": bl_no}, "name")
		if existing:
			shipments.append(
				frappe.db.get_value(
					"Shipment", {"container": existing, "shipment_type": "Own Goods (Trading)"}, "name"
				)
			)
			skipped.append(f"already imported: {container_no} ({bl_no})")
			continue

		container = frappe.get_doc(
			{
				"doctype": "Container",
				"direction": "Import",
				"container_no": container_no,
				"bl_no": bl_no,
				"branch": branch,
				"eta": row.get("eta"),
				"ata": row.get("date_received") if row.get("status") == "Arrived" else None,
				"notes": row.get("comment"),
			}
		)
		# The manifest belongs to the box (Model B). These are the company's own
		# goods, so the lines carry no customer tag.
		for pkg in row["packages"]:
			container.append(
				"contents",
				{
					"item": ensure_item(pkg.get("description"), "Import"),
					"description": pkg.get("description"),
					"qty": pkg.get("qty") or 0,
					"unit": pkg.get("unit") or "Nos",
					"weight_kg": pkg.get("weight_kg"),
					"declared_value": pkg.get("declared_value"),
				},
			)
		container.flags.ignore_permissions = True
		container.insert(ignore_permissions=True)
		created["containers"] += 1

		if row.get("qty_needs_split"):
			first = row["packages"][0]
			skipped.append(
				f"{container_no}: qty {first['qty']} put on '{first['description']}' — split it manually"
			)

		notes = []
		if row.get("supplier"):
			notes.append(f"Supplier: {row['supplier']}")
		if row.get("invoice_ref"):
			notes.append(f"Supplier invoice: {row['invoice_ref']}")
		if row.get("comment"):
			notes.append(row["comment"])

		shipment = frappe.get_doc(
			{
				"doctype": "Shipment",
				"shipment_type": "Own Goods (Trading)",
				"direction": "Import",
				"containers": [{"container": container.name}],
				"branch": branch,
				"notes": "\n".join(notes) or None,
			}
		)
		shipment.flags.ignore_permissions = True
		shipment.insert(ignore_permissions=True)
		shipments.append(shipment.name)
		created["shipments"] += 1

		milestone = STATUS_MILESTONE.get((row.get("status") or "").strip())
		if milestone:
			frappe.get_doc(
				{
					"doctype": "Tracking Event",
					"container": container.name,
					"milestone": milestone,
					"event_datetime": row.get("date_received") or row.get("eta") or frappe.utils.now_datetime(),
					"source": "Manual",
					"notify": 0,
				}
			).insert(ignore_permissions=True)
			created["events"] += 1

	manifests = manifests_for([s for s in shipments if s])
	for row in data["distributions"]:
		# Find the trading shipment whose manifest matches this product — their
		# sheet shortens names ("Hen Leg Quarter" vs "US Hen Leg Quarter").
		target, package_desc, package_item = None, None, None
		for ship_name in shipments:
			if not ship_name:
				continue
			for line in manifests.get(ship_name, []):
				if _norm(row["product"]) in _norm(line["description"]) and flt(line["qty"]) > 0:
					target, package_desc, package_item = ship_name, line["description"], line["item"]
					break
			if target:
				break
		if not target:
			skipped.append(f"distribution '{row['product']}' → no matching arrived shipment; enter manually")
			continue

		dup = frappe.db.exists(
			"Distribution Entry",
			{
				"shipment": target,
				"recipient": row["recipient"],
				"qty": flt(row["qty"]),
				"delivery_date": row.get("delivery_date"),
			},
		)
		if dup:
			skipped.append(f"distribution already imported: {row['recipient']} / {row['product']} / {row['qty']}")
			continue

		frappe.get_doc(
			{
				"doctype": "Distribution Entry",
				"shipment": target,
				"item": package_item,
				"product": package_desc,
				"qty": flt(row["qty"]),
				"recipient": row["recipient"],
				"destination": row.get("destination"),
				"unit_price": row.get("unit_price"),
				"delivery_date": row.get("delivery_date"),
			}
		).insert(ignore_permissions=True)
		created["distributions"] += 1

	frappe.db.commit()
	print(
		f"Imported: {created['containers']} containers, {created['shipments']} shipments, "
		f"{created['events']} events, {created['distributions']} distributions"
	)
	for s in skipped:
		print(f"  note: {s}")
	return created


def load_opening():
	"""Load the opening balance committed in the repo (used by the patch)."""
	path = opening_data_path()
	if not os.path.exists(path):
		print(f"No opening data at {path} — nothing to import")
		return {}
	with open(path) as f:
		return load(json.load(f))


def run(path=None):
	"""Import straight from a spreadsheet on this machine."""
	try:
		data = parse_workbook(path)
	except FileNotFoundError as e:
		print(f"File not found: {e}")
		return
	return load(data)
